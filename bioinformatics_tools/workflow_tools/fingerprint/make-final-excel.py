#!/usr/bin/env python3
"""make-final-excel.py — convert FINAL-scored-labeled-genes-annotated.tsv to a colored Excel workbook.

Whole-row reviewer coloring:
  - needs_review == "yes"  -> red gradient by confidence tier
      (low=pink -> highest=strong red)
  - needs_review != "yes"  -> green gradient by confidence tier
      (low=very light green -> highest=stronger bright green)
  - non-coding / unscored  -> light gray

Text color is auto-picked for contrast on each shade.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

csv.field_size_limit(10_000_000)

# ── FINAL-file coloring — MATCHES the operon-diagram figures
#    (reportfig_lib.CONF_TIER_COLOR). Each row gets a LIGHT tint of its
#    confidence-tier colour; the key "answer" cells (tier, adjusted-confidence,
#    needs-review, operon-context direction) get the BRIGHT figure colour with
#    auto-contrast text. Kept in sync with api/routers/ssh.py. ─────────────────
_TIER_BRIGHT = {
    "highest": "1F77FF",   # blue
    "high":    "00B84D",   # green
    "medium":  "FFCC00",   # yellow
    "fair":    "FF8C00",   # orange
    "low":     "EE2233",   # red
}
_CTX_RAISED = "6B8E23"      # operon context increases confidence (olive)
_CTX_LOWERED = "8B0000"     # decreases (dark red)
_REVIEW_YES = "EE2233"      # needs review (bright red)
_REVIEW_NO = "1E9E57"       # ok (green)
_ROW_NONCODING_BG = "F2F2F2"
_ROW_NONCODING_FG = "8A8A8A"
# Whole-row highlight by feature type / EC availability (added for transparency):
_ROW_NONCDS_BG = "FFF2A8"   # light yellow: non-CDS features (rna / prophage)
_ROW_NOEC_BG = "F1A9A0"     # light red: CDS genes with NO EC evidence (not an enzyme)


def _hex_lum(h: str) -> float:
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255.0


def _contrast_fg(h: str) -> str:
    return "000000" if _hex_lum(h) > 0.55 else "FFFFFF"


def _tint_hex(h: str, toward_white: float = 0.86) -> str:
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = round(r + (255 - r) * toward_white)
    g = round(g + (255 - g) * toward_white)
    b = round(b + (255 - b) * toward_white)
    return f"{r:02X}{g:02X}{b:02X}"

# ── header ─────────────────────────────────────────────────────────────────
_HDR_BG = "203864"
_HDR_FG = "FFFFFF"

# ── column widths (characters) ──────────────────────────────────────────────
_COL_WIDTHS: dict[str, int] = {
    "organism_name":   35,
    "feature_id":      30,
    "gene_id":         28,
    "na_seq":           8,
    "aa_seq":           8,
    "na_length":       10,
    "aa_length":       10,
    "gene_start":      12,
    "gene_end":        12,
    "RAST_feature_type": 16,
    "RAST_strand":     10,
    "operon_id":       16,
    "operon_member_count": 14,
    "operon_gene_position_in_operon": 20,
    "gene_fingerprint_with_hash": 40,
    "best_consensus_product_descriptor": 36,
    "product_descriptor_source":    20,
    "product_descriptor_source_id": 24,
    "product_descriptor_hierarchy": 22,
    "product_descriptor_audit_trail": 36,
    "product_descriptor_confirmatory_summary_specialized_tools": 38,
    "product_descriptor_hierarchy_tier_name": 32,
    "c1_score_database_coverage": 14,
    "c1_score_reasoning": 40,
    "c2_score_operon_probability": 14,
    "c2_score_reasoning": 40,
    "c3_score_operon_context": 14,
    "c3_reasoning": 40,
    "c4_score_EC_agreement": 14,
    "c4_reasoning": 40,
    "c4_ec_agreement_status": 20,
    "preliminary_confidence_c1_c4": 16,
    "final_confidence_operon_context": 18,
    "does_context_improve_confidence?": 20,
    "confidence_tier": 14,
    "needs_review": 12,
    "needs_review_reason": 44,
    "gene_id_copy": 28,
    "gene_start_copy": 12,
    "gene_end_copy": 12,
    "best_consensus_product_descriptor_copy": 36,
}
_DEFAULT_WIDTH = 16


_FILL_CACHE: dict[str, PatternFill] = {}
_FONT_CACHE: dict[tuple[str, bool], Font] = {}


def _fill(hex_color: str) -> PatternFill:
    # reuse one PatternFill object per colour -- openpyxl then dedups by identity,
    # which turns ~200k per-cell style assignments from minutes into seconds.
    f = _FILL_CACHE.get(hex_color)
    if f is None:
        f = PatternFill("solid", fgColor=hex_color)
        _FILL_CACHE[hex_color] = f
    return f


def _font(hex_color: str, bold: bool = False) -> Font:
    key = (hex_color, bold)
    f = _FONT_CACHE.get(key)
    if f is None:
        f = Font(color=hex_color, bold=bold)
        _FONT_CACHE[key] = f
    return f


def _safe_float(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _norm_col(name: str) -> str:
    # FINAL_ANNOTATION_WITH_CONFIDENCE supports prefixed headers like:
    #   "[AN]-NEEDS_REVIEW?"  (legacy)
    #   "Column-AN: NEEDS_REVIEW?"  (current)
    return re.sub(r"^(?:\[[A-Z]+\]-|Column-[A-Z]+:\s*)", "", str(name or "").strip(), flags=re.IGNORECASE).strip().lower()


def _row_get(row: dict, *candidate_names: str) -> str:
    normalized_targets = {_norm_col(n) for n in candidate_names}
    for k, v in row.items():
        if _norm_col(k) in normalized_targets:
            return v
    return ""


def _row_tint(row: dict) -> tuple[str, str]:
    """(bg, fg) whole-row highlight, priority: non-CDS features (light yellow) >
    CDS with no EC evidence (light red) > confidence-tier tint. Bright accent
    cells are added on top per-column."""
    ftype = str(_row_get(row, "FEATURE_TYPE", "feature_type")).strip().lower()
    if ftype and ftype != "cds":                         # rna, prophage, ...
        return _ROW_NONCDS_BG, "000000"
    final = _safe_float(_row_get(
        row,
        "final_confidence_operon_context",
        "ADJUSTED_CONFIDENCE_WITH_OPERON_CONTEXT",
    ))
    if final is None:
        return _ROW_NONCODING_BG, _ROW_NONCODING_FG
    ec_status = str(_row_get(row, "EC_EVIDENCE_STATUS", "c4_ec_agreement_status")).strip().lower()
    if ec_status == "no_evidence":                       # coding, but no enzyme call to verify
        return _ROW_NOEC_BG, "000000"
    tier = str(_row_get(row, "confidence_tier", "CONFIDENCE_TIER")).strip().lower()
    bright = _TIER_BRIGHT.get(tier, _TIER_BRIGHT["medium"])
    # deeper tint (was 0.86 -> near-white) so the WHOLE confidence-tier row reads as
    # one colour band for easy row tracking, while the bright accent cells still pop
    return _tint_hex(bright, 0.72), "000000"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input",  required=True,
                        help="FINAL-scored-labeled-genes-annotated.tsv")
    parser.add_argument("--output", required=True,
                        help="Output .xlsx path")
    args = parser.parse_args()

    in_path  = Path(args.input)
    out_path = Path(args.output)

    if not in_path.is_file():
        print(f"[make-final-excel] ERROR: input not found: {in_path}", file=sys.stderr)
        raise SystemExit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation Results"

    with open(in_path, newline="") as fh:
        reader = csv.DictReader(fh, delimiter="\t")
        headers = list(reader.fieldnames or [])

        # ── header row ──────────────────────────────────────────────────────
        hdr_fill = _fill(_HDR_BG)
        hdr_font = _font(_HDR_FG, bold=True)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        # normalized column-name -> 1-based index (headers are "Column-XX: name")
        norm_idx: dict[str, int] = {}
        for i, h in enumerate(headers, 1):
            norm_idx.setdefault(_norm_col(h), i)

        def cidx(*names: str) -> int | None:
            for nm in names:
                j = norm_idx.get(_norm_col(nm))
                if j:
                    return j
            return None

        tier_i = cidx("confidence_tier", "CONFIDENCE_TIER")
        tierh_i = cidx("confidence_tier_hybrid", "CONFIDENCE_TIER_hybrid")
        adj_i = cidx("final_confidence_operon_context", "ADJUSTED_CONFIDENCE_WITH_OPERON_CONTEXT")
        adjh_i = cidx("ADJUSTED_CONFIDENCE_WITH_OPERON_CONTEXT_hybrid")
        rev_i = cidx("needs_review?", "NEEDS_REVIEW?", "needs_review")
        ctx_i = cidx("does_operon_context_improve_confidence?", "DOES_OPERON_CONTEXT_IMPROVE_CONFIDENCE?")

        def _accent(cell_ri: int, col_i: int | None, bg: str | None, bold: bool = True) -> None:
            if not col_i or not bg:
                return
            cell = ws.cell(row=cell_ri, column=col_i)
            cell.fill = _fill(bg)
            cell.font = _font(_contrast_fg(bg), bold=bold)

        # ── data rows ───────────────────────────────────────────────────────
        # ONE shared alignment object for every data cell (creating a fresh
        # Alignment per cell is what made a 4.6k-row sheet take minutes).
        data_align = Alignment(wrap_text=False, vertical="top")
        n = 0
        flag_counts = {"noncds_yellow": 0, "noec_red": 0, "tier": 0,
                       "review_yes": 0}
        for ri, row in enumerate(reader, 2):
            row_bg, row_fg = _row_tint(row)
            row_fill = _fill(row_bg)
            row_font = _font(row_fg)

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(h, ""))
                cell.alignment = data_align
                cell.fill = row_fill
                cell.font = row_font

            # bright accent cells (match the operon-diagram figures)
            tier_adj = str(_row_get(row, "confidence_tier", "CONFIDENCE_TIER")).strip().lower()
            tier_hyb = str(_row_get(row, "confidence_tier_hybrid", "CONFIDENCE_TIER_hybrid")).strip().lower()
            _accent(ri, tier_i, _TIER_BRIGHT.get(tier_adj))
            _accent(ri, adj_i, _TIER_BRIGHT.get(tier_adj))
            _accent(ri, tierh_i, _TIER_BRIGHT.get(tier_hyb))
            _accent(ri, adjh_i, _TIER_BRIGHT.get(tier_hyb))
            rev = str(_row_get(row, "needs_review?", "NEEDS_REVIEW?", "needs_review")).strip().lower()
            if rev == "yes":
                _accent(ri, rev_i, _REVIEW_YES)
            elif rev == "no":
                _accent(ri, rev_i, _REVIEW_NO)
            ctx = str(_row_get(row, "does_operon_context_improve_confidence?",
                               "DOES_OPERON_CONTEXT_IMPROVE_CONFIDENCE?")).strip().lower()
            if "increase" in ctx:
                _accent(ri, ctx_i, _CTX_RAISED, bold=False)
            elif "decrease" in ctx:
                _accent(ri, ctx_i, _CTX_LOWERED, bold=False)

            # tally by the whole-row highlight category (see _row_tint)
            if row_bg == _ROW_NONCDS_BG:
                flag_counts["noncds_yellow"] += 1
            elif row_bg == _ROW_NOEC_BG:
                flag_counts["noec_red"] += 1
            else:
                flag_counts["tier"] += 1
            if rev == "yes":
                flag_counts["review_yes"] += 1
            n += 1

    # ── freeze panes and row height ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40

    # ── column widths ───────────────────────────────────────────────────────
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTHS.get(h, _DEFAULT_WIDTH)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"[make-final-excel] {n} genes → {out_path}")
    print(f"  {len(headers)} columns; whole-row highlighting applied")
    print(f"    light-yellow (non-CDS: rna/prophage) : {flag_counts['noncds_yellow']}")
    print(f"    light-red    (CDS, no EC evidence)   : {flag_counts['noec_red']}")
    print(f"    tier-tinted  (CDS with EC evidence)  : {flag_counts['tier']}")
    print(f"    (of which needs_review=yes hard flag : {flag_counts['review_yes']})")


if __name__ == "__main__":
    main()
