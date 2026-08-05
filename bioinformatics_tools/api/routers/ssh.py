"""
SSH and job management endpoints.

Thin routing layer — delegates to job_store, job_runner, and ssh utilities.

All endpoints (except /health) require a valid Bearer token. The token is
validated by get_current_user(), which returns the user's cluster credentials.
_build_connection() decrypts the stored private key and builds a per-user
SSHConnection for each request.
"""
import io
import json
import logging
import posixpath
import re
import shlex
import threading
import uuid
from datetime import datetime, timezone

import pandas as pd
import yaml
from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl.styles import Border, Font, PatternFill, Side
from fastapi.responses import Response, StreamingResponse

from dataclasses import asdict

from bioinformatics_tools.api.auth import decrypt_private_key, get_current_user
from bioinformatics_tools.api.models import GenomeSend, SlurmSend
from bioinformatics_tools.api.services import job_history_client, job_runner
from bioinformatics_tools.api.services.job_store import job_store
from bioinformatics_tools.utilities import ssh_sftp, ssh_slurm
from bioinformatics_tools.utilities.ssh_connection import make_user_connection
from bioinformatics_tools.workflow_tools.workflow_helpers import GENOME_EXTENSIONS
from bioinformatics_tools.workflow_tools.workflow_registry import (
    MARGIE_SB_PHASED_TOOLS,
    WORKFLOWS,
    REQUIRED_SYSTEM_PARAMS,
    workflow_path_params,
)

LOGGER = logging.getLogger(__name__)

router = APIRouter(prefix="/v1/ssh", tags=["ssh"])

# Workflows visible on the frontend but not yet implemented.
STUB_WORKFLOWS: set[str] = {"custom_microbiome"}

# (job_id, path) -> (mtime, size, total_lines). Wiped on API restart, same
# as job_store's in-memory state. Avoids re-running wc -l (a full file
# scan) on every page click for a file whose content hasn't changed --
# output files don't change once a job completes.
_line_count_cache: dict[tuple[str, str], tuple[float, int, int]] = {}


def _validate_relative_path(path: str, *, label: str = "file") -> None:
    """Raise HTTPException(400) if path attempts directory traversal.

    Shared by job_files (subdir), download_file (path), and view_file
    (path) -- all three take a user-supplied relative path under a job's
    work_dir.
    """
    if path and (path.startswith("/") or ".." in path.split("/")):
        raise HTTPException(status_code=400, detail=f"Invalid {label} path")


def _cfg_get(cfg: dict, key: str, default=None):
    """Get a nested config value using dot notation, returning default if any
    segment is missing."""
    value = cfg
    for part in key.split('.'):
        if isinstance(value, dict) and part in value:
            value = value[part]
        else:
            return default
    return value


def _cfg_set(cfg: dict, key: str, value) -> None:
    """Set a nested config value using dot notation, creating parents."""
    parts = key.split('.')
    target = cfg
    for part in parts[:-1]:
        next_value = target.get(part)
        if not isinstance(next_value, dict):
            next_value = {}
            target[part] = next_value
        target = next_value
    target[parts[-1]] = value


def _first_nonempty(*values):
    for value in values:
        if value is None:
            continue
        if str(value).strip() == "":
            continue
        return value
    return None


def _expand_remote_home(path: str, home_dir: str) -> str:
    if path.startswith("~"):
        return path.replace("~", home_dir, 1)
    return path


def _is_user_scoped_db(path: str, username: str) -> bool:
    """Legacy heuristic: DB/path basename starts with username- prefix."""
    return posixpath.basename(path).startswith(f"{username}-")


def _owner_marker_path(path: str, *, is_dir: bool) -> str:
    """Companion marker path storing ownership/provenance metadata."""
    if is_dir:
        return f"{path.rstrip('/')}/.margie-owner.json"
    return f"{path}.margie-owner.json"


def _read_owner_marker(conn, path: str, *, is_dir: bool) -> dict | None:
    """Read ownership marker JSON for a path, if it exists and is valid."""
    marker = _owner_marker_path(path, is_dir=is_dir)
    cmd = f"if [ -f {shlex.quote(marker)} ]; then cat {shlex.quote(marker)}; fi"
    exit_code, output = _run_remote_check(conn, cmd)
    if exit_code != 0 or not output:
        return None
    try:
        data = json.loads(output)
        return data if isinstance(data, dict) else None
    except json.JSONDecodeError:
        LOGGER.warning("Ignoring invalid ownership marker at %s", marker)
        return None


def _write_owner_marker(current_user: dict, conn, path: str, *, is_dir: bool,
                        source_path: str | None = None) -> None:
    """Write ownership marker for a user-scoped promoted path."""
    marker = _owner_marker_path(path, is_dir=is_dir)
    payload = {
        "scope": "user",
        "owner_username": current_user["username"],
        "owner_cluster_username": current_user["cluster_username"],
        "kind": "directory" if is_dir else "file",
        "path": path,
    }
    if source_path:
        payload["source_path"] = source_path
    ssh_sftp.write_remote_text_file(
        marker,
        json.dumps(payload, indent=2, sort_keys=True) + "\n",
        connection=conn,
    )


def _classify_path_scope(current_user: dict, conn, path: str, *, is_dir: bool) -> str:
    """Classify path as user/shared using marker-first logic.

    Returns one of:
    - 'user'   : explicitly owned by current user
    - 'shared' : no marker and no user prefix

    Raises HTTPException if marker exists but belongs to a different user.
    """
    marker = _read_owner_marker(conn, path, is_dir=is_dir)
    if marker is not None:
        owner = marker.get("owner_username")
        scope = marker.get("scope")
        if scope == "user" and owner == current_user["username"]:
            return "user"
        if scope == "user" and owner and owner != current_user["username"]:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Configured path '{path}' is marked as private to user '{owner}'. "
                    "Please select a shared template path or your own private path."
                ),
            )

    # Backward-compatible fallback for pre-marker paths.
    return "user" if _is_user_scoped_db(path, current_user["username"]) else "shared"


def _versioned_user_db_path(template_db: str, username: str, version: int) -> str:
    """Build '<dir>/<username>-<stem>-vN<ext>' from a shared template DB path."""
    directory = posixpath.dirname(template_db)
    filename = posixpath.basename(template_db)
    stem, ext = posixpath.splitext(filename)
    target_name = f"{username}-{stem}-v{version}{ext}"
    return posixpath.join(directory, target_name) if directory else target_name


def _versioned_user_dir_path(template_dir: str, username: str, version: int) -> str:
    """Build '<dir>/<username>-<name>-vN' from a shared directory path."""
    parent = posixpath.dirname(template_dir.rstrip('/'))
    name = posixpath.basename(template_dir.rstrip('/'))
    target_name = f"{username}-{name}-v{version}"
    return posixpath.join(parent, target_name) if parent else target_name


def _find_existing_user_db_versions(conn, template_db: str, username: str) -> list[int]:
    """List existing version numbers for username-prefixed copies of template_db."""
    directory = posixpath.dirname(template_db)
    if not directory:
        directory = "."
    filename = posixpath.basename(template_db)
    stem, ext = posixpath.splitext(filename)
    prefix = f"{username}-{stem}-v"

    try:
        entries = ssh_sftp.list_remote_dir(directory, connection=conn)
    except FileNotFoundError:
        return []
    except Exception:
        return []

    versions: list[int] = []
    for entry in entries:
        if entry.get("type") != "file":
            continue
        name = entry.get("name") or ""
        if not name.startswith(prefix) or not name.endswith(ext):
            continue
        middle = name[len(prefix):]
        if ext:
            middle = middle[:-len(ext)]
        if middle.isdigit():
            versions.append(int(middle))
    return sorted(versions)


def _find_existing_user_dir_versions(conn, template_dir: str, username: str) -> list[int]:
    """List existing version numbers for username-prefixed copies of a directory."""
    parent = posixpath.dirname(template_dir.rstrip('/'))
    if not parent:
        parent = "."
    base = posixpath.basename(template_dir.rstrip('/'))
    prefix = f"{username}-{base}-v"

    try:
        entries = ssh_sftp.list_remote_dir(parent, connection=conn)
    except FileNotFoundError:
        return []
    except Exception:
        return []

    versions: list[int] = []
    for entry in entries:
        if entry.get("type") != "directory":
            continue
        name = entry.get("name") or ""
        if not name.startswith(prefix):
            continue
        suffix = name[len(prefix):]
        if suffix.isdigit():
            versions.append(int(suffix))
    return sorted(versions)


def _promote_shared_file_to_user_file(current_user: dict, conn, raw_path: str) -> tuple[str, bool]:
    """Resolve a writable per-user file path, copying shared template on first use."""
    expanded = _expand_remote_home(raw_path, current_user["home_dir"])
    username = current_user["username"]

    if _classify_path_scope(current_user, conn, expanded, is_dir=False) == "user":
        # Upgrade legacy user-prefixed paths by backfilling marker metadata.
        if _read_owner_marker(conn, expanded, is_dir=False) is None:
            _write_owner_marker(current_user, conn, expanded, is_dir=False)
        return expanded, False

    try:
        ssh_sftp.check_remote_file(expanded, connection=conn)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Shared file path does not exist or is unreadable: '{expanded}'. Details: {exc}",
        )

    existing_versions = _find_existing_user_db_versions(conn, expanded, username)
    if existing_versions:
        target = _versioned_user_db_path(expanded, username, existing_versions[-1])
        if _read_owner_marker(conn, target, is_dir=False) is None:
            _write_owner_marker(current_user, conn, target, is_dir=False, source_path=expanded)
    else:
        target = _versioned_user_db_path(expanded, username, 1)
        target_dir = posixpath.dirname(target) or "."
        cmd = (
            f"mkdir -p {shlex.quote(target_dir)} && "
            f"cp {shlex.quote(expanded)} {shlex.quote(target)}"
        )
        exit_code, output = _run_remote_check(conn, cmd)
        if exit_code != 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Could not create user-specific file from shared template. "
                    f"Source: '{expanded}', target: '{target}'. Details: {output or 'copy failed'}"
                ),
            )
        LOGGER.info("Created user-specific shared file copy for %s: %s", username, target)
        _write_owner_marker(current_user, conn, target, is_dir=False, source_path=expanded)
    return target, True


def _promote_shared_dir_to_user_dir(current_user: dict, conn, raw_path: str) -> tuple[str, bool]:
    """Resolve a writable per-user directory path, creating versioned copy dir on first use."""
    expanded = _expand_remote_home(raw_path, current_user["home_dir"])
    username = current_user["username"]

    if _classify_path_scope(current_user, conn, expanded.rstrip('/'), is_dir=True) == "user":
        cmd = f"mkdir -p {shlex.quote(expanded)}"
        exit_code, output = _run_remote_check(conn, cmd)
        if exit_code != 0:
            raise HTTPException(
                status_code=400,
                detail=f"Could not ensure user directory exists: '{expanded}'. Details: {output or 'mkdir failed'}",
            )
        if _read_owner_marker(conn, expanded, is_dir=True) is None:
            _write_owner_marker(current_user, conn, expanded, is_dir=True)
        return expanded, False

    existing_versions = _find_existing_user_dir_versions(conn, expanded, username)
    if existing_versions:
        target = _versioned_user_dir_path(expanded, username, existing_versions[-1])
    else:
        target = _versioned_user_dir_path(expanded, username, 1)

    cmd = f"mkdir -p {shlex.quote(target)}"
    exit_code, output = _run_remote_check(conn, cmd)
    if exit_code != 0:
        raise HTTPException(
            status_code=400,
            detail=f"Could not create user-specific directory: '{target}'. Details: {output or 'mkdir failed'}",
        )
    if not existing_versions:
        LOGGER.info("Created user-specific shared directory for %s: %s", username, target)
    if _read_owner_marker(conn, target, is_dir=True) is None:
        _write_owner_marker(current_user, conn, target, is_dir=True, source_path=expanded)
    return target, True


def _promote_shared_main_db_to_user_db(current_user: dict, user_config: dict, conn) -> tuple[str, bool]:
    """Resolve a writable per-user main_database path.

    If main_database already points to a username-prefixed file, keep it.
    Otherwise treat it as a shared template and switch to a user-specific
    versioned copy alongside it (reuse highest existing version if present,
    else create v1 by copying the template).

    Returns: (resolved_main_db_path, config_changed)
    """
    raw = user_config.get("main_database")
    if not raw or str(raw).strip() == "":
        raise HTTPException(
            status_code=400,
            detail="main_database is not configured. Please configure it in your Profile settings.",
        )

    username = current_user["username"]
    expanded = _expand_remote_home(str(raw).strip(), current_user["home_dir"])
    if _classify_path_scope(current_user, conn, expanded, is_dir=False) == "user":
        if _read_owner_marker(conn, expanded, is_dir=False) is None:
            _write_owner_marker(current_user, conn, expanded, is_dir=False)
        return expanded, False

    # Shared template mode: switch to user-specific path in same directory.
    try:
        ssh_sftp.check_remote_file(expanded, connection=conn)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Configured main_database does not exist or is unreadable: '{expanded}'. Details: {exc}",
        )

    existing_versions = _find_existing_user_db_versions(conn, expanded, username)
    if existing_versions:
        target = _versioned_user_db_path(expanded, username, existing_versions[-1])
        if _read_owner_marker(conn, target, is_dir=False) is None:
            _write_owner_marker(current_user, conn, target, is_dir=False, source_path=expanded)
    else:
        target = _versioned_user_db_path(expanded, username, 1)
        target_dir = posixpath.dirname(target) or "."
        cmd = (
            f"mkdir -p {shlex.quote(target_dir)} && "
            f"cp {shlex.quote(expanded)} {shlex.quote(target)}"
        )
        exit_code, output = _run_remote_check(conn, cmd)
        if exit_code != 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Could not create user-specific main_database from shared template. "
                    f"Source: '{expanded}', target: '{target}'. Details: {output or 'copy failed'}"
                ),
            )
        LOGGER.info("Created user-specific main_database for %s: %s", username, target)
        _write_owner_marker(current_user, conn, target, is_dir=False, source_path=expanded)

    user_config["main_database"] = target
    return target, True


def _resolve_effective_main_db(current_user: dict, conn, user_config: dict, *, persist: bool) -> str:
    """Resolve main_database and optionally persist config updates.

    During workflow launches we persist the promoted user-scoped DB path so
    subsequent runs keep using it by default.
    """
    main_db, changed = _promote_shared_main_db_to_user_db(current_user, user_config, conn)
    if changed and persist:
        ssh_sftp.write_remote_yaml(_config_path(current_user["home_dir"]), user_config, connection=conn)
    return main_db


def _resolve_effective_margie_sb_writable_paths(current_user: dict, conn, user_config: dict, *, persist: bool) -> dict:
    """Promote writable margie_sb shared paths to per-user versioned paths.

    Returns mapping of config key -> resolved path.
    """
    resolved: dict[str, str] = {}
    changed_any = False

    file_keys = (
        (
            'margie_sb.operon_database.occ_reference_pkl',
            'operon_database.occ_reference_pkl',
            '/depot/lindems/data/margie/operon-database/occ_reference.pkl',
        ),
        (
            'margie_sb.fingerprint_database.path',
            'fingerprint_database.path',
            '/depot/lindems/data/margie/fingerprint-database/fingerprint-database.tsv',
        ),
    )
    dir_keys = (
        (
            'margie_sb.genome_pool.path',
            'genome_pool.path',
            '/depot/lindems/data/margie/genome-pool',
        ),
        (
            'margie_sb.scoring_results_historical.path',
            'scoring_results_historical.path',
            '/depot/lindems/data/margie/scoring-results-historical',
        ),
        (
            'margie_sb.final_tables_depot.path',
            'final_tables_depot.path',
            '/depot/lindems/data/margie/final-tables',
        ),
        (
            'margie_sb.sqlite_pipeline_snapshot.path',
            'sqlite_pipeline_snapshot.path',
            '/depot/lindems/data/margie/sqlite/pipeline-version',
        ),
    )

    for key, legacy_key, default_value in file_keys:
        raw = _first_nonempty(_cfg_get(user_config, key), _cfg_get(user_config, legacy_key), default_value)
        target, changed = _promote_shared_file_to_user_file(current_user, conn, str(raw))
        _cfg_set(user_config, key, target)
        resolved[key] = target
        changed_any = changed_any or changed

    for key, legacy_key, default_value in dir_keys:
        raw = _first_nonempty(_cfg_get(user_config, key), _cfg_get(user_config, legacy_key), default_value)
        target, changed = _promote_shared_dir_to_user_dir(current_user, conn, str(raw))
        _cfg_set(user_config, key, target)
        resolved[key] = target
        changed_any = changed_any or changed

    if changed_any and persist:
        ssh_sftp.write_remote_yaml(_config_path(current_user["home_dir"]), user_config, connection=conn)

    return resolved


def _run_remote_check(conn, command: str) -> tuple[int, str]:
    ssh = conn.connect()
    _, stdout, stderr = ssh.exec_command(command)
    exit_code = stdout.channel.recv_exit_status()
    output = (stdout.read().decode(errors="replace") + stderr.read().decode(errors="replace")).strip()
    return exit_code, output


def _assert_remote_writable(conn, path: str, *, label: str, treat_as_file: bool = False) -> None:
    target_dir = posixpath.dirname(path) if treat_as_file else path
    if not target_dir:
        target_dir = "."
    probe = posixpath.join(target_dir, f".margie_write_test_{uuid.uuid4().hex}")
    command = (
        f"mkdir -p {shlex.quote(target_dir)} && "
        f"test -w {shlex.quote(target_dir)} && "
        f"touch {shlex.quote(probe)} && rm -f {shlex.quote(probe)}"
    )
    exit_code, output = _run_remote_check(conn, command)
    if exit_code != 0:
        raise HTTPException(
            status_code=400,
            detail=f"{label} is not writable or cannot be created at '{path}'. "
                   f"Please update your Profile config path settings. Details: {output or 'permission/path check failed'}",
        )


def _validate_margie_sb_shared_paths(user_config: dict, conn, home_dir: str) -> None:
    """Fail early if MARGIE_SB shared storage paths are not writable.

    New namespaced keys are preferred; legacy top-level keys remain supported
    as fallback for older configs.
    """
    writable_paths = [
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.operon_database.occ_reference_pkl'),
                _cfg_get(user_config, 'operon_database.occ_reference_pkl'),
                '/depot/lindems/data/margie/operon-database/occ_reference.pkl',
            ),
            'margie_sb.operon_database.occ_reference_pkl',
            True,
        ),
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.fingerprint_database.path'),
                _cfg_get(user_config, 'fingerprint_database.path'),
                '/depot/lindems/data/margie/fingerprint-database/fingerprint-database.tsv',
            ),
            'margie_sb.fingerprint_database.path',
            True,
        ),
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.genome_pool.path'),
                _cfg_get(user_config, 'genome_pool.path'),
                '/depot/lindems/data/margie/genome-pool',
            ),
            'margie_sb.genome_pool.path',
            False,
        ),
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.scoring_results_historical.path'),
                _cfg_get(user_config, 'scoring_results_historical.path'),
                '/depot/lindems/data/margie/scoring-results-historical',
            ),
            'margie_sb.scoring_results_historical.path',
            False,
        ),
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.final_tables_depot.path'),
                _cfg_get(user_config, 'final_tables_depot.path'),
                '/depot/lindems/data/margie/final-tables',
            ),
            'margie_sb.final_tables_depot.path',
            False,
        ),
        (
            _first_nonempty(
                _cfg_get(user_config, 'margie_sb.sqlite_pipeline_snapshot.path'),
                _cfg_get(user_config, 'sqlite_pipeline_snapshot.path'),
                '/depot/lindems/data/margie/sqlite/pipeline-version',
            ),
            'margie_sb.sqlite_pipeline_snapshot.path',
            False,
        ),
    ]

    for raw_path, key_name, treat_as_file in writable_paths:
        expanded = _expand_remote_home(str(raw_path), home_dir)
        _assert_remote_writable(conn, expanded, label=f"Shared path '{key_name}'", treat_as_file=treat_as_file)


def _resolve_job_work_dir(job_id: str, current_user: dict, conn) -> str:
    """Resolves a job's work_dir, falling back to persistent history if the
    job isn't in the in-memory job_store (e.g. after a dane-api restart) --
    the same fallback get_job_status already uses, applied here so file
    browsing/download/view work for resumed/historical jobs too, not just
    job_status itself.

    Raises HTTPException(404) if the job can't be found anywhere (live or
    history), or HTTPException(400) if found but has no work_dir yet.
    """
    job = job_store.get(job_id)
    if job is not None:
        if job.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        work_dir = job.get("work_dir")
    else:
        try:
            user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
        except Exception:
            raise HTTPException(status_code=404, detail="Job not found")

        main_db = user_config.get('main_database')
        row = (
            job_history_client.get_job(
                conn,
                main_db,
                job_id,
                owner_username=current_user["username"],
                owner_cluster_username=current_user["cluster_username"],
            )
            if main_db else None
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Job not found")
        work_dir = row.get("work_dir")

    if not work_dir:
        raise HTTPException(status_code=400, detail="No working directory available for this job")
    return work_dir


_SCORE_TIER_COLORS: dict[str, str] = {
    "highest":          "1a9641",
    "high":             "a6d96a",
    "moderate":         "ffffbf",
    "fair":             "fdae61",
    "low":              "d7191c",
}

_CONFIDENCE_TIER_COLORS: dict[str, str] = {
    "high":             "1a9641",
    "moderate":         "ffffbf",
    "low":              "d7191c",
    "flagged_for_review": "9e1985",
}

# ACS tier: yellow (low) → dark blue (highest), matching make-final-excel.py
_ACS_TIER_ROW_COLORS: dict[str, str] = {
    "low":      "FFFDE7",
    "fair":     "FFF3E0",
    "moderate": "E8F4FC",
    "high":     "D6E8F7",
    "highest":  "C5DDEF",
    "NOT_APPLICABLE_NON_CODING": "F5F5F5",
}

# White text on dark backgrounds, black text on light ones.
_TIER_FONT_COLORS: dict[str, str] = {
    "1a9641": "FFFFFF",
    "a6d96a": "000000",
    "ffffbf": "000000",
    "fdae61": "000000",
    "d7191c": "FFFFFF",
    "9e1985": "FFFFFF",
    "FFFDE7": "000000",
    "FFF3E0": "000000",
    "E8F4FC": "000000",
    "D6E8F7": "000000",
    "C5DDEF": "000000",
    "F5F5F5": "888888",
}


# ── FINAL publication file coloring — MATCHES the operon-diagram FIGURES
#    (reportfig_lib.CONF_TIER_COLOR). Every row is tinted edge to edge with its
#    CONFIDENCE_TIER_HYBRID tier colour so the sheet is scannable; review rows
#    additionally get a box border. Kept in sync with
#    workflow_tools/fingerprint/make-final-excel.py. ───────────────────────────
_TIER_BRIGHT = {
    "highest": "1F77FF",   # blue
    "high":    "00B84D",   # green
    "medium":  "FFCC00",   # yellow
    "fair":    "FF8C00",   # orange
    "low":     "EE2233",   # red
}
_ROW_NONCODING_TINT = "F2F2F2"
_ROW_NONCODING_FG = "8A8A8A"


def _tint_hex(h: str, toward_white: float = 0.86) -> str:
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = round(r + (255 - r) * toward_white)
    g = round(g + (255 - g) * toward_white)
    b = round(b + (255 - b) * toward_white)
    return f"{r:02X}{g:02X}{b:02X}"


def _norm_col(name: str) -> str:
    # FINAL_ANNOTATION_WITH_CONFIDENCE supports prefixed headers like:
    #   "[AN]-NEEDS_REVIEW?"  (legacy)
    #   "Column-AN: NEEDS_REVIEW?"  (current)
    return re.sub(r"^(?:\[[A-Z]+\]-|Column-[A-Z]+:\s*)", "", str(name or "").strip(), flags=re.IGNORECASE).strip().lower()


def _series_get(row: pd.Series, *candidate_names: str) -> str:
    targets = {_norm_col(n) for n in candidate_names}
    for key in row.index:
        if _norm_col(key) in targets:
            return row.get(key, "")
    return ""


def _has_any_column(df: pd.DataFrame, *candidate_names: str) -> bool:
    targets = {_norm_col(n) for n in candidate_names}
    for col in df.columns:
        if _norm_col(col) in targets:
            return True
    return False


_ROW_TINT = 0.72                                     # tier-colour lightness per row
_REVIEW_SIDE = Side(style="medium", color="000000")  # box border on review rows


def _row_tint(row: pd.Series) -> tuple[str, str]:
    """(bg, fg) whole-row colour = the row's CONFIDENCE_TIER_HYBRID tier colour,
    tinted and applied across the entire row (matches make-final-excel.py). Rows
    with no scored tier -- empty or NOT_APPLICABLE_NON_CODING (rna / prophage) --
    get grey. This is the only colouring; no accents."""
    tier = str(_series_get(row, "confidence_tier_hybrid", "CONFIDENCE_TIER_hybrid")).strip().lower()
    if tier not in _TIER_BRIGHT:
        return _ROW_NONCODING_TINT, _ROW_NONCODING_FG
    return _tint_hex(_TIER_BRIGHT[tier], _ROW_TINT), "000000"


def _apply_review_flag_colors(ws, df: pd.DataFrame) -> None:
    """Colour each data row edge to edge with its CONFIDENCE_TIER_HYBRID tier
    colour; rows flagged NEEDS_REVIEW? = yes get a box border around the whole
    row. Matches make-final-excel.py (no per-cell accents)."""
    n_cols = len(df.columns)
    for offset, (_, row) in enumerate(df.iterrows()):
        r = offset + 2  # row 1 is the header
        bg, fg = _row_tint(row)
        fill = PatternFill(fill_type="solid", fgColor=bg)
        font = Font(color=fg)
        review = str(_series_get(row, "needs_review?", "NEEDS_REVIEW?", "needs_review")).strip().lower() == "yes"
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = font
            if review:                          # box border around the whole review row
                cell.border = Border(
                    top=_REVIEW_SIDE, bottom=_REVIEW_SIDE,
                    left=_REVIEW_SIDE if col == 1 else None,
                    right=_REVIEW_SIDE if col == n_cols else None)


def _apply_tier_row_colors(ws, df: pd.DataFrame) -> None:
    """Color each data row. For the FINAL publication file (which carries the
    two-stage confidence columns) use review-flag coloring; otherwise fall
    back to per-tier row coloring when a tier column exists."""
    if _has_any_column(df, "final_confidence_operon_context", "ADJUSTED_CONFIDENCE_WITH_OPERON_CONTEXT"):
        _apply_review_flag_colors(ws, df)
        return
    if "ACS_tier" in df.columns:
        tier_col, color_map = "ACS_tier", _ACS_TIER_ROW_COLORS
    elif "confidence_score_tier" in df.columns:
        tier_col, color_map = "confidence_score_tier", _SCORE_TIER_COLORS
    elif "confidence_tier" in df.columns:
        tier_col, color_map = "confidence_tier", _CONFIDENCE_TIER_COLORS
    else:
        return
    n_cols = len(df.columns)
    for row_idx, tier_val in enumerate(df[tier_col], start=2):  # row 1 is the header
        hex_color = color_map.get(str(tier_val).strip())
        if hex_color is None:
            continue
        fill = PatternFill(fill_type="solid", fgColor=hex_color)
        font_color = _TIER_FONT_COLORS.get(hex_color, "000000")
        font = Font(color=font_color)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.font = font


def _detect_delimiter(path: str, header: str) -> str:
    """Pick a column delimiter: by extension first, else sniff the header.

    Naive -- no RFC4180 quote-handling. Every sampled real output file is
    quote-free TSV; a CSV with delimiters embedded in quoted fields would
    misparse. Acceptable v1 limitation, not silently papered over.
    """
    lower = path.lower()
    if lower.endswith(".csv"):
        return ","
    if lower.endswith(".tsv"):
        return "\t"
    return "\t" if "\t" in header else ","


def _get_available_workflows() -> list[dict]:
    """
    Build the list of available workflows from WORKFLOWS registry.
    Returns detailed metadata for each workflow including tools, params, etc.
    Automatically merges REQUIRED_SYSTEM_PARAMS with workflow-specific params.
    """
    workflows = []

    # Add workflows from WORKFLOWS registry
    for wf_id, wf_key in WORKFLOWS.items():
        # Skip internal test workflows
        if wf_id in ['example', 'selftest']:
            continue

        # Convert dataclass to dict and add computed fields
        wf_dict = asdict(wf_key)
        wf_dict['id'] = wf_key.cmd_identifier
        wf_dict['containers'] = [{'name': sif[0], 'version': sif[1]} for sif in wf_key.sif_files]

        # Merges system-wide required params, this workflow's own root-path settings,
        # and the workflow's other params. System params come first since they're
        # infra-level. input_path/output_path apply to every workflow; sif_path and
        # db_root are only included when this workflow actually has a local-folder
        # sif lookup / a unified db_root fallback to point at (see
        # workflow_path_params()'s docstring for why that distinction matters).
        path_params = workflow_path_params(
            wf_id,
            include_sif=wf_key.local_sif_only,
            include_db_root=wf_key.supports_db_root,
            supports_batch_input=wf_key.supports_batch_input,
        )
        wf_dict['configurable_params'] = REQUIRED_SYSTEM_PARAMS + path_params + (wf_key.configurable_params or [])

        workflows.append(wf_dict)

    # Adds stub workflows (not yet implemented but visible)
    # Even stub workflows get system params since they'll need them when implemented
    workflows.append({
        'id': 'custom_microbiome',
        'label': 'Custom Microbiome',
        'description': 'Custom microbiome annotation workflow (coming soon)',
        'full_description': 'A specialized workflow for microbiome annotation. This workflow is currently under development.',
        'tools': [],
        'configurable_params': REQUIRED_SYSTEM_PARAMS,  # Stub still needs system params
        'database_deps': [],
        'docs_url': None,
        'containers': [],
        'cmd_identifier': 'custom_microbiome',
        'snakemake_file': '',
        'other': [],
        'sif_files': [],
    })

    return workflows


def _build_connection(current_user: dict):
    """Decrypt the user's stored private key and return a ready SSHConnection."""
    private_key = decrypt_private_key(current_user['private_key_encrypted'])
    return make_user_connection(
        current_user['cluster_host'],
        current_user['cluster_username'],
        private_key,
    )


def _config_path(home_dir: str) -> str:
    """Remote path to the user's BSP config file."""
    return f'{home_dir}/.config/bioinformatics-tools/config.yaml'


def _yaml_block(data: dict, indent: int = 0) -> str:
    block = yaml.safe_dump(data, sort_keys=False, default_flow_style=False).rstrip()
    prefix = ' ' * indent
    return '\n'.join(f'{prefix}{line}' if line else '' for line in block.splitlines())


def _set_nested_value(target: dict, parts: list[str], value):
    current = target
    for part in parts[:-1]:
        current = current.setdefault(part, {})
    current[parts[-1]] = value


def _ordered_workflow_params(workflow_id: str, params: list[dict]) -> list[dict]:
    if workflow_id != 'margie_sb':
        return list(params)

    shared_group_order = {
        'operon_database': 0,
        'fingerprint_database': 1,
        'genome_pool': 2,
        'scoring_results_historical': 3,
        'final_tables_depot': 4,
        'sqlite_pipeline_snapshot': 5,
        'report_figures': 6,
    }
    tool_phase_order = {tool['key']: (tool['phase'], index) for index, tool in enumerate(MARGIE_SB_PHASED_TOOLS)}

    def phase_leaf_order(leaf: str) -> int:
        return {
            'partition': 0,
            'max_parallel_genomes': 1,
            'max_parallel_tools': 2,
            'threads': 3,
            'mem_mb': 4,
            'runtime': 5,
            'db': 6,
            'sif': 7,
        }.get(leaf, 99)

    def sort_key(param: dict) -> tuple:
        parts = param['param'].split('.')
        if len(parts) == 2 and parts[1] in {'default_threads', 'default_mem_mb', 'default_runtime'}:
            return (0, 0, 0, parts[1])
        if len(parts) >= 3 and parts[1].startswith('phase'):
            phase_num = int(parts[1][5:]) if parts[1][5:].isdigit() else 999
            return (1, phase_num, phase_leaf_order(parts[-1]), param['param'])
        if len(parts) >= 3 and parts[1] in shared_group_order:
            return (2, shared_group_order[parts[1]], phase_leaf_order(parts[-1]), param['param'])
        if len(parts) >= 2 and parts[0] == 'margie_sb':
            phase_num, tool_index = tool_phase_order.get(parts[1], (999, 999))
            return (3, phase_num, tool_index, phase_leaf_order(parts[-1]), param['param'])
        return (4, param['param'])

    return sorted(params, key=sort_key)


def _default_params_for_workflow(workflow_id: str, workflow) -> list[dict]:
    """Return the per-workflow params that should be materialized in a
    default config payload.

    Keeps this aligned with /workflows metadata shown in Profile: root-path
    params are injected based on workflow capabilities, then merged with the
    workflow's own configurable params.
    """
    path_params = workflow_path_params(
        workflow_id,
        include_sif=workflow.local_sif_only,
        include_db_root=workflow.supports_db_root,
        supports_batch_input=workflow.supports_batch_input,
    )
    combined = path_params + (workflow.configurable_params or [])

    # De-duplicate by key while preserving first-seen order.
    seen: set[str] = set()
    unique: list[dict] = []
    for param in combined:
        key = param.get('param')
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(param)
    return unique


def _build_default_config_payload() -> dict:
    config: dict = {
        'main_database': '~/.local/share/bioinformatics-tools/my-db.db',
        'compute': {'cluster_default': {}},
    }

    for param in REQUIRED_SYSTEM_PARAMS:
        if param['param'].startswith('compute.cluster_default.'):
            key = param['param'].split('.')[-1]
            default_value = param.get('default')
            config['compute']['cluster_default'][key] = default_value if default_value is not None else ''

    for workflow_id, workflow in WORKFLOWS.items():
        params = _default_params_for_workflow(workflow_id, workflow)
        if not params:
            continue

        section: dict = {}
        for param in _ordered_workflow_params(workflow_id, params):
            parts = param['param'].split('.')

            # Params are usually namespaced (e.g. "margie_sb.sif_path").
            # Strip that prefix so values live under one workflow block:
            # margie_sb:
            #   sif_path: ...
            if parts and parts[0] == workflow_id:
                parts = parts[1:]
            if not parts:
                continue

            default_value = param.get('default')
            if default_value is not None:
                _set_nested_value(section, parts, default_value)

        if section:
            config[workflow_id] = section

    return config


def _build_default_config_text(config: dict) -> str:
    return yaml.safe_dump(config, sort_keys=False, default_flow_style=False)


@router.get("/workflows")
def list_workflows(current_user: dict = Depends(get_current_user)):
    """Return the list of user-facing workflows with detailed metadata."""
    return _get_available_workflows()


@router.get("/health")
def health_check():
    """Test endpoint to verify API is working. No auth required."""
    return {"status": "success"}


@router.get("/status")
def ssh_status(current_user: dict = Depends(get_current_user)):
    """Check whether the BSP server can reach the user's cluster via SSH.

    Returns 200 either way -- the UI polls this and a red banner is a better
    answer than a failed request. But it now reports WHY, and whether the user
    can do anything about it. Previously every failure collapsed to
    {"connected": false} with the reason visible only in the server log, so an
    undecryptable stored key was indistinguishable from an unreachable cluster.
    That cost real debugging time: the backend knew exactly what was wrong and
    said so in its log while the UI just showed "no connection".
    """
    try:
        conn = _build_connection(current_user)
        ssh = conn.connect()
        # Do NOT close it: the client is pooled and shared across requests.
        return {"connected": True, "host": current_user["cluster_host"]}
    except HTTPException as exc:
        # _build_connection -> decrypt_private_key raises this when the stored
        # key cannot be decrypted, which happens when BSP_ENCRYPTION_KEY has been
        # regenerated since the account was created. Fernet is authenticated
        # encryption, so the key material is unrecoverable -- re-registering is
        # the only fix, and the user needs telling that plainly.
        detail = str(getattr(exc, "detail", exc))
        undecryptable = "decrypt" in detail.lower()
        LOGGER.warning("SSH status check failed for user %s: %s",
                       current_user["username"], detail)
        return {
            "connected": False,
            "host": current_user["cluster_host"],
            "reason": "key_undecryptable" if undecryptable else "error",
            "detail": (
                "Your stored SSH key cannot be decrypted, because the server's "
                "encryption key changed after this account was created. The key "
                "cannot be recovered — please register a new account to continue."
                if undecryptable else detail
            ),
            "action": "re-register" if undecryptable else None,
        }
    except Exception as exc:
        LOGGER.warning("SSH status check failed for user %s: %s",
                       current_user["username"], exc)
        return {
            "connected": False,
            "host": current_user["cluster_host"],
            "reason": "unreachable",
            "detail": f"Could not reach {current_user['cluster_host']}: {exc}",
            "action": None,
        }


@router.get("/config")
def get_config(current_user: dict = Depends(get_current_user)):
    """Read the user's ~/.config/bioinformatics-tools/config.yaml from their cluster via SFTP."""
    conn = _build_connection(current_user)
    path = _config_path(current_user["home_dir"])
    try:
        data = ssh_sftp.read_remote_yaml(path, connection=conn)
        return data
    except Exception as exc:
        LOGGER.error("Failed to read remote config for %s: %s", current_user["username"], exc)
        raise HTTPException(status_code=500, detail=f"Failed to read remote config: {exc}")


@router.put("/config")
def save_config(config: dict, current_user: dict = Depends(get_current_user)):
    """Write a config dict back to the user's cluster as YAML via SFTP."""
    conn = _build_connection(current_user)
    path = _config_path(current_user["home_dir"])
    try:
        ssh_sftp.write_remote_yaml(path, config, connection=conn)
        return {"success": True}
    except Exception as exc:
        LOGGER.error("Failed to write remote config for %s: %s", current_user["username"], exc)
        raise HTTPException(status_code=500, detail=f"Failed to write remote config: {exc}")


@router.post("/config/create-default")
def create_default_config(current_user: dict = Depends(get_current_user)):
    """Create a default config file with all system defaults populated."""
    conn = _build_connection(current_user)
    path = _config_path(current_user["home_dir"])

    default_config = _build_default_config_payload()
    default_config_text = _build_default_config_text(default_config)

    try:
        ssh_sftp.write_remote_text_file(path, default_config_text, connection=conn)
        LOGGER.info("Created default config for user %s at %s", current_user["username"], path)
        return {"success": True, "config": default_config}
    except Exception as exc:
        LOGGER.error("Failed to create default config for %s: %s", current_user["username"], exc)
        raise HTTPException(status_code=500, detail=f"Failed to create default config: {exc}")


@router.post("/test-path-writable")
def test_path_writable(path_data: dict, current_user: dict = Depends(get_current_user)):
    """Test if a path on the cluster is writable by attempting to create parent directories and a test file."""
    conn = _build_connection(current_user)
    test_path = path_data.get("path", "").strip()

    if not test_path:
        raise HTTPException(status_code=400, detail="Path is required")

    try:
        ssh = conn.connect()

        # Expand ~ to actual home directory
        if test_path.startswith("~"):
            test_path = test_path.replace("~", current_user["home_dir"], 1)

        # Get the directory (remove filename if present)
        import posixpath
        test_dir = posixpath.dirname(test_path)

        # Try to create the directory structure
        _, stdout, stderr = ssh.exec_command(f'mkdir -p "{test_dir}" 2>&1 && echo "DIR_OK"')
        output = stdout.read().decode().strip()

        if "DIR_OK" not in output:
            pass  # pooled client: closing it would break concurrent requests (see SSHConnection pool)
            return {
                "writable": False,
                "error": f"Cannot create directory: {test_dir}",
                "details": output
            }

        # Try to write a test file
        test_file = f"{test_path}.write_test"
        _, stdout, stderr = ssh.exec_command(f'touch "{test_file}" 2>&1 && rm -f "{test_file}" 2>&1 && echo "WRITE_OK"')
        output = stdout.read().decode().strip()

        pass  # pooled client: closing it would break concurrent requests (see SSHConnection pool)

        if "WRITE_OK" in output:
            return {"writable": True}
        else:
            return {
                "writable": False,
                "error": f"Path is not writable: {test_path}",
                "details": output
            }

    except Exception as exc:
        LOGGER.error("Failed to test path writability for %s: %s", current_user["username"], exc)
        return {
            "writable": False,
            "error": f"Failed to test path: {str(exc)}"
        }


@router.post("/run_slurm")
def run_slurm(content: SlurmSend, current_user: dict = Depends(get_current_user)):
    """Submit a SLURM job and return the job ID immediately."""
    conn = _build_connection(current_user)
    job_id = ssh_slurm.submit_slurm_job(script_content=content.script, connection=conn)
    return {"success": True, "job_id": job_id, "message": "Job submitted successfully"}


@router.post("/run_ssh")
def run_ssh(content: SlurmSend, current_user: dict = Depends(get_current_user)):
    """Execute an SSH command and return output."""
    LOGGER.info('Running run_ssh for user %s', current_user["username"])
    conn = _build_connection(current_user)
    std_txt = ssh_slurm.submit_ssh_job(cmd=content.script, connection=conn)
    return {"success": True, "std_txt": std_txt, "message": "Job submitted successfully"}


def _check_genome_path_exists(genome_path: str, workflow: str, conn) -> None:
    """Raise HTTPException(400) if genome_path doesn't exist on the cluster,
    or (for batch-input workflows) the folder has no recognized genome file.

    Shared by run_workflow (fresh submissions) and resume_job/restart_job
    (re-validating a previously-known-good path, since the remote file
    could have been deleted or moved since the original run)."""
    wf_key = WORKFLOWS.get(workflow)
    supports_batch_input = bool(wf_key and wf_key.supports_batch_input)
    try:
        if supports_batch_input:
            attr = ssh_sftp.check_remote_path_kind(genome_path, conn)
            if attr == 'directory':
                entries = ssh_sftp.list_remote_dir(genome_path, conn)
                has_genome_file = any(
                    e['type'] == 'file' and e['name'].lower().endswith(GENOME_EXTENSIONS)
                    for e in entries
                )
                if not has_genome_file:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No recognized genome files (e.g. .fasta, .fa, .fna) found in folder: '{genome_path}'",
                    )
        else:
            ssh_sftp.check_remote_file(genome_path, conn)
    except FileNotFoundError:
        raise HTTPException(
            status_code=400,
            detail=f"Path not found on the cluster: '{genome_path}'. "
                   "Make sure the path is a Negishi path, not a path on your local machine.",
        )
    except IsADirectoryError:
        raise HTTPException(
            status_code=400,
            detail=f"Path points to a directory, not a file: '{genome_path}'",
        )
    except HTTPException:
        raise
    except Exception as exc:
        LOGGER.warning("File pre-check failed for %s: %s", genome_path, exc)
        raise HTTPException(
            status_code=400,
            detail=f"Could not verify path on cluster: {exc}",
        )


def _launch_job(
    *, genome_path: str, workflow: str, base_output_dir: str,
    selected_tools: list[str] | None, current_user: dict, conn, main_db: str | None,
    slurm_account: str | None = None,
    slurm_partition: str | None = None,
    relaunched_from: str | None = None,
    copy_from_work_dir: str | None = None,
    run_full_operon_map: bool = False,
) -> dict:
    """Shared job-launch sequence: generate job_id/timestamp/output_dir,
    optionally copy a previous run's output_dir into the new one first
    (Resume), persist to job_store, build the dane_wf command (adding
    margie_sb.resume: true when copy_from_work_dir is set, so Snakemake's
    mtime-only rerun triggers recognize the copied-forward outputs as
    already done -- see workflow_tools/workflow.py's build_executable), and
    submit.

    Does NOT do workflow-id/stub/config/genome-path pre-flight validation --
    callers (run_workflow, resume_job, restart_job) each do whatever subset
    of that is appropriate for their entry point before calling this.
    """
    selected_tools_csv = ",".join(selected_tools) if selected_tools is not None else None
    selected_tools_arg = f" {workflow}.selected_tools: {selected_tools_csv}" if selected_tools_csv else ""
    # Opt-in full-genome operon atlas: top-level flag the smk gate reads
    # (rc_bool('run_full_operon_map', False)). Runs downstream of the report figures.
    full_operon_map_arg = " run_full_operon_map: true" if run_full_operon_map else ""

    job_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime('%Y-%m-%d-%H%M')
    output_dir = f"{base_output_dir.rstrip('/')}/{timestamp}"

    if copy_from_work_dir:
        ssh_sftp.copy_remote_directory(copy_from_work_dir, output_dir, connection=conn)
        try:
            ssh_sftp.rewrite_path_references(output_dir, copy_from_work_dir, output_dir, connection=conn)
        except Exception as exc:
            # Cosmetic cleanup only (confirmed: nothing downstream reads the
            # stale provenance columns this fixes up) -- never let a failure
            # here block the resumed job from launching.
            LOGGER.warning("Could not rewrite stale path references for resumed job: %s", exc)

    job_store.create(
        job_id, genome_path, user_id=current_user["user_id"],
        workflow=workflow, output_dir=output_dir,
        selected_tools=selected_tools_csv, relaunched_from=relaunched_from,
        persist_owner_username=current_user["username"],
        persist_owner_cluster_username=current_user["cluster_username"],
        persist_db_path=main_db, persist_connection=conn,
    )
    job_store.update(job_id, work_dir=output_dir)

    # The CLI dispatcher (caragols) matches do_<a>_<b> against the SEPARATE
    # tokens "<a> <b>", not the underscore-joined string -- e.g. do_margie_sb
    # is invoked as "margie sb", not "margie_sb". workflow itself (and every
    # config key built from it, e.g. selected_tools_arg/resume_arg) stays
    # underscore-joined, matching the registry's cmd_identifier.
    dispatch_tokens = workflow.replace('_', ' ')
    resume_arg = f" {workflow}.resume: true" if copy_from_work_dir else ""
    # Licensing: acceptance was verified in run_workflow (server-side). Pass it
    # (and the user's entitlement) to the CLI so its own gate does not re-prompt
    # this non-interactive run, and so it disables the same tools the user isn't
    # licensed for. See workflow_tools/license_gate.py.
    from bioinformatics_tools.api import licensing
    _lic_ent = licensing.get_entitlement(current_user["username"])
    _lic_csv = ",".join(_lic_ent.get("licensed_tools") or [])
    license_env = (
        f"MARGIE_LICENSE_ACCEPTED='{licensing.load_terms()['version']}' "
        f"MARGIE_USAGE_TYPE='{_lic_ent.get('usage_type') or ''}' "
        f"MARGIE_LICENSED_TOOLS='{_lic_csv}' "
    )
    # Invokes dane_wf directly from ~/bioinformatics-tools/.venv (an editable
    # install -- code changes there are picked up instantly, no reinstall
    # needed) rather than through `uvx --from`. uvx re-resolves/caches the
    # local package on its own schedule, independent of whether
    # ~/bioinformatics-tools (see _ensure_remote_deployment_symlink in
    # api/main.py) actually points at fresh code -- confirmed adding 3+
    # minutes per run and still serving a stale build without --refresh.
    # The venv binary has neither problem: ~0.4s overhead, always current.
    command = (
        f"{license_env}~/bioinformatics-tools/.venv/bin/dane_wf {dispatch_tokens}"
        f" input: {genome_path} output_dir: {output_dir}{selected_tools_arg}{full_operon_map_arg}{resume_arg}"
    )
    job_runner.submit_job(
        job_id,
        command,
        connection=conn,
        driver_account=slurm_account,
        driver_partition=slurm_partition,
    )

    return {"success": True, "job_id": job_id, "output_dir": output_dir, "message": "Job submitted successfully"}


@router.post("/run_workflow")
def run_workflow(genome_data: GenomeSend, current_user: dict = Depends(get_current_user)):
    """Submit a genome analysis workflow by name."""
    # License gate (server-side enforcement — the analyze page also gates in the
    # UI, but a run must never proceed without a recorded acceptance of the
    # current licensing terms).
    from bioinformatics_tools.api import licensing
    if not licensing.has_accepted_current_terms(current_user["username"]):
        raise HTTPException(
            status_code=403,
            detail="You must accept the current MARGIE licensing terms before running an analysis.",
        )

    available_workflows = _get_available_workflows()
    allowed_ids = {wf["id"] for wf in available_workflows}

    if genome_data.workflow not in allowed_ids:
        raise HTTPException(status_code=400, detail=f"Unknown workflow '{genome_data.workflow}'. Available: {sorted(allowed_ids)}")

    if genome_data.workflow in STUB_WORKFLOWS:
        raise HTTPException(status_code=501, detail=f"Workflow '{genome_data.workflow}' is not yet implemented. Check back soon!")

    conn = _build_connection(current_user)

    # Pre-flight: validate required config values are set
    config_path = _config_path(current_user["home_dir"])
    try:
        user_config = ssh_sftp.read_remote_yaml(config_path, connection=conn)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Configuration file not found. Please create a configuration in your Profile settings first."
        )

    # Validate required fields
    missing_fields = []

    # Check main_database
    main_db = user_config.get('main_database')
    if not main_db or str(main_db).strip() == '':
        missing_fields.append('main_database')

    # Check compute.cluster_default.account
    account = user_config.get('compute', {}).get('cluster_default', {}).get('account')
    if not account or str(account).strip() == '':
        missing_fields.append('compute.cluster_default.account (SLURM account)')
    slurm_account = str(account).strip() if account else None
    partition = user_config.get('compute', {}).get('cluster_default', {}).get('partition')
    slurm_partition = str(partition).strip() if partition else None

    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Required configuration missing: {', '.join(missing_fields)}. "
                   "Please configure these in your Profile settings before running workflows."
        )

    if genome_data.workflow == 'margie_sb':
        _resolve_effective_margie_sb_writable_paths(current_user, conn, user_config, persist=True)
        _validate_margie_sb_shared_paths(user_config, conn, current_user["home_dir"])

    # Shared template DB -> per-user DB promotion (persisted), so concurrent
    # users stop writing to one SQLite file.
    main_db = _resolve_effective_main_db(current_user, conn, user_config, persist=True)

    # Resolves genome path / output dir, falling back to the user's global config
    # defaults (input_path / output_path) when the request didn't specify one.
    genome_path = genome_data.genome_path or user_config.get(genome_data.workflow, {}).get('input_path')
    if not genome_path or str(genome_path).strip() == '':
        raise HTTPException(
            status_code=400,
            detail="No genome file or folder specified, and no input_path default is configured. "
                   "Set one in your Profile settings or pass genome_path explicitly.",
        )

    _check_genome_path_exists(genome_path, genome_data.workflow, conn)

    # Validates phase/tool selection, if given -- catch typos here rather than
    # have them silently no-op as an unrecognized run_<tool> config key.
    # `is not None` (not a truthiness check) matters here: an explicit empty
    # list means "run nothing", which must NOT be treated the same as
    # omitting the field entirely ("run everything").
    if genome_data.selected_tools is not None:
        if not genome_data.selected_tools:
            raise HTTPException(
                status_code=400,
                detail="selected_tools was empty -- select at least one tool/phase to run, "
                       "or omit the field entirely to run everything.",
            )
        valid_tool_keys = {tool['key'] for tool in MARGIE_SB_PHASED_TOOLS}
        unknown = set(genome_data.selected_tools) - valid_tool_keys
        if unknown:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown tool key(s) in selected_tools: {sorted(unknown)}. "
                       f"Available: {sorted(valid_tool_keys)}",
            )
        # Refuse tools the user is not licensed for (mirrors the CLI gate and the
        # greyed-out tools in the analyze UI -- defence in depth).
        _ent = licensing.get_entitlement(current_user["username"])
        _disabled = licensing.disabled_tool_ids(
            _ent.get("usage_type"), _ent.get("licensed_tools")
        ) & valid_tool_keys
        _blocked = set(genome_data.selected_tools) & _disabled
        if _blocked:
            raise HTTPException(
                status_code=403,
                detail=f"These tools require a license you have not recorded: {sorted(_blocked)}. "
                       "Update your usage type / licensed tools when accepting the terms "
                       "(Profile), or remove them from your selection.",
            )

    base_dir = (genome_data.output_dir or user_config.get(genome_data.workflow, {}).get('output_path') or current_user['home_dir']).rstrip('/')

    # run_full_operon_map: enable if EITHER the request asks for it OR it is
    # persisted in the user's per-workflow config (Profile settings). The saved
    # config is the reliable channel -- it reaches the backend server-side and
    # does not depend on the analysis-page checkbox making it into the payload.
    saved_full_operon_map = bool(
        user_config.get(genome_data.workflow, {}).get('run_full_operon_map', False))
    effective_full_operon_map = bool(genome_data.run_full_operon_map) or saved_full_operon_map

    return _launch_job(
        genome_path=genome_path, workflow=genome_data.workflow,
        base_output_dir=base_dir, selected_tools=genome_data.selected_tools,
        current_user=current_user, conn=conn, main_db=main_db,
        slurm_account=slurm_account, slurm_partition=slurm_partition,
        run_full_operon_map=effective_full_operon_map,
    )


def _job_from_history_row(row: dict) -> dict:
    """Shape a persisted api_jobs row like an in-memory job_store entry, so
    the front-end's job page can render it the same way whether the job is
    still live or was resumed after a dane-api restart.

    logs/slurm_jobs/containers are only ever a final snapshot, taken once
    by job_store.finalize() at job completion/failure -- never persisted
    incrementally (that would mean an SSH round-trip per log line), so a
    job that's still mid-run when dane-api restarts has none of this yet,
    and a job that died without ever reaching finalize() (e.g. dane-api
    itself crashed) never gets one at all. sub_jobs/report/progress/
    steps_done/total remain pure live-session detail, never persisted.
    job_history_client already JSON-decodes slurm_jobs/containers back
    into real lists before this row ever reaches here.
    """
    return {
        "job_id": row["job_id"],
        "owner_username": row.get("owner_username"),
        "owner_cluster_username": row.get("owner_cluster_username"),
        "status": row["status"],
        "phase": row.get("phase"),
        "genome_path": row.get("genome_path"),
        "workflow": row.get("workflow"),
        "work_dir": row.get("work_dir"),
        "selected_tools": row.get("selected_tools"),
        "relaunched_from": row.get("relaunched_from"),
        "start_time": row.get("created_at"),
        "sub_jobs": [],
        "slurm_jobs": row.get("slurm_jobs") or [],
        "containers": row.get("containers") or [],
        "logs": row.get("logs") or "",
        "resumed_from_history": True,
    }

def _load_job_for_action(job_id: str, current_user: dict, conn) -> dict:
    """Resolve a job_id to enough info to relaunch it (genome_path, workflow,
    work_dir, selected_tools, status), whether it's still live in job_store
    or only in persisted history. Raises 404/403 the same way get_job_status
    does.

    Deliberately separate from get_job_status itself: that endpoint also
    does SLURM-reconciliation (still_active/status_note) this lookup
    doesn't need, and from _resolve_job_work_dir (file-serving endpoints
    only need work_dir; resume/restart need the full launch-relevant
    shape). Minor duplication of the job_store-then-history-fallback
    pattern across these three is accepted -- collapsing them would couple
    endpoints with different response-shape needs."""
    job = job_store.get(job_id)
    if job is not None:
        if job.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        return job

    try:
        user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
    except Exception:
        raise HTTPException(status_code=404, detail="Job not found")
    main_db = user_config.get('main_database')
    row = (
        job_history_client.get_job(
            conn,
            main_db,
            job_id,
            owner_username=current_user["username"],
            owner_cluster_username=current_user["cluster_username"],
        )
        if main_db else None
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return _job_from_history_row(row)


_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
_POTENTIALLY_STALE_STATUSES = {"pending", "running", "snakemake"}

# A run that is still being submitted (status "running", phase "Submitting via
# SSH") has no .rc sentinel, no driver process yet, and no SLURM jobs -- which
# is byte-for-byte the same signature the reconcilers otherwise read as
# "interrupted / driver stopped". That race mislabelled brand-new jobs as
# cancelled on the history page while they were still on their way to the
# cluster. Give a freshly-touched row this grace window before ever calling it
# dead: submission has time to spawn the driver and queue its SLURM jobs, at
# which point the normal alive-checks keep it running on their own merits.
_SUBMIT_GRACE_SECONDS = 300


def _row_age_seconds(row: dict) -> float | None:
    """Seconds since this history row was last touched, or None if it has no
    parseable timestamp. Prefers updated_at (bumped when the row entered the
    submitting state) and falls back to created_at."""
    stamp = row.get("updated_at") or row.get("created_at")
    if not stamp:
        return None
    try:
        ts = datetime.fromisoformat(stamp)
    except (TypeError, ValueError):
        return None
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - ts).total_seconds()


def _within_submit_grace(row: dict) -> bool:
    """True while a row is too young to be judged interrupted -- it is almost
    certainly still submitting (no driver/SLURM visible yet)."""
    age = _row_age_seconds(row)
    return age is not None and age < _SUBMIT_GRACE_SECONDS
# States that mean "this job still occupies the cluster". COMPLETING belongs
# here: a job tearing down is emphatically not finished, and leaving it out
# made a workflow read as dead whenever a poll happened to land while its
# whole current batch was in that state -- which for short rules is most of
# the time. CONFIGURING/COMPLETED_* are the other transient squeue states.
_ACTIVE_SLURM_STATES = ("RUNNING", "PENDING", "COMPLETING", "CONFIGURING",
                        "RESIZING", "SUSPENDED", "REQUEUED")

# Guards the rehydrate-and-reattach below: the job page polls every 10s and
# FastAPI runs sync endpoints in a threadpool, so without this two overlapping
# polls both see an empty job_store and both start a watcher on the same run.
_REATTACH_LOCK = threading.Lock()


def _try_reattach(job_id: str, row: dict, conn, main_db: str | None,
                  current_user: dict) -> dict | None:
    """Take a still-running job back over after a dane-api restart.

    A workflow run is detached (setsid + nohup), so it survives the API that
    started it. Its in-memory job_store entry does not. Everything needed to
    resume watching is still on the cluster -- the log under
    ~/.local/share/bsp/jobs/<job_id>.log and the .rc exit sentinel beside it --
    and the tail has always replayed from line 1, so re-reading it re-derives
    every SLURM job, container and progress line the dead session had parsed.

    Returns the rehydrated live job dict, or None if this run cannot or should
    not be reattached (already finished, or another thread got there first).
    """
    # Checked BEFORE claiming the job, because `tail -F` on a log that will
    # never grow never returns, and would park one of the runner's four
    # workers for the lifetime of the process.
    #
    # Replay is right in exactly two cases: the run is still going (its log is
    # still being written), or it finished while the API was down (exit
    # sentinel present) -- in which case the replay recovers the whole run AND
    # its real exit code, instead of leaving a finished job stuck at "running"
    # forever. A run with neither was interrupted; the squeue path below
    # reports that instead. See ssh_slurm.is_replayable.
    try:
        probe = ssh_slurm.probe_run(job_id, connection=conn)
    except Exception as exc:
        LOGGER.warning("Could not probe job %s for reattach: %s", job_id, exc)
        return None
    if not ssh_slurm.is_replayable(probe):
        LOGGER.info("Job %s is not replayable (%s); not reattaching", job_id, probe)
        return None

    with _REATTACH_LOCK:
        if job_store.exists(job_id):
            return job_store.get(job_id)          # another poll won the race

        job_store.create(
            job_id, row.get("genome_path") or "", user_id=current_user["user_id"],
            workflow=row.get("workflow"), output_dir=row.get("work_dir"),
            selected_tools=row.get("selected_tools"),
            relaunched_from=row.get("relaunched_from"),
        )
        # create() would re-INSERT a history row that already exists, so
        # persistence is attached afterwards instead: subsequent status/phase
        # changes still reach the user's history, no duplicate row is written.
        job_store.attach_persistence(job_id, main_db, conn)
        # Carry the row's own status/phase over. create() starts every job at
        # "pending", and this one is not pending -- it has been running since
        # before the restart. Without this the first poll after a restart
        # would report a live job as pending, and the front-end would drop the
        # Emergency Stop button for it.
        job_store.update(job_id, work_dir=row.get("work_dir"),
                         status=row.get("status") or "running",
                         phase=row.get("phase") or "Reattaching to running job")

    LOGGER.info("Reattaching to job %s after an API restart", job_id)
    job_runner.submit_job(job_id, command="", connection=conn, reattach=True)
    return job_store.get(job_id)


@router.get("/job_status/{job_id}")
def get_job_status(
    job_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get status of a running job. Falls back to persistent history (e.g.
    after a dane-api restart wiped the in-memory job_store) before giving
    up. Returns 403 if the in-memory job belongs to a different user --
    history rows can't make that check since they live in each user's own
    main_database already, with no cross-user data to separate.

    For a non-terminal history row (nothing live is tracking it anymore,
    so its persisted status could be stale), also checks squeue to see
    whether it's genuinely still active on the cluster -- adds
    still_active/status_note to the response, purely additive, so this
    never changes the existing status field or breaks older clients."""
    job = job_store.get(job_id)
    if job is not None:
        if job.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        return {**job, "cluster_host": current_user["cluster_host"]}

    conn = _build_connection(current_user)
    try:
        user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
    except Exception:
        raise HTTPException(status_code=404, detail="Job not found")

    main_db = user_config.get('main_database')
    row = (
        job_history_client.get_job(
            conn,
            main_db,
            job_id,
            owner_username=current_user["username"],
            owner_cluster_username=current_user["cluster_username"],
        )
        if main_db else None
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Job not found")

    # A non-terminal row means this run was still going when dane-api lost
    # track of it. Take it back over rather than settling for a squeue
    # snapshot: replaying its log restores the logs, SLURM jobs, containers
    # and progress the dead session had parsed, and from here on it updates
    # live again. The squeue path below stays as the fallback for runs with
    # no replayable log.
    if row["status"] not in _TERMINAL_STATUSES:
        live = _try_reattach(job_id, row, conn, main_db, current_user)
        if live is not None:
            return {**live, "cluster_host": current_user["cluster_host"]}

    result = {**_job_from_history_row(row), "cluster_host": current_user["cluster_host"]}

    if row["status"] not in _TERMINAL_STATUSES and row.get("work_dir"):
        try:
            matches = ssh_slurm.find_active_jobs_in_workdir(
                row["work_dir"], current_user["cluster_username"], connection=conn,
            )
        except Exception as exc:
            LOGGER.warning("SLURM reconciliation check failed for job %s: %s", job_id, exc)
            matches = []
        active_matches = [m for m in matches if m["state"] in _ACTIVE_SLURM_STATES]
        still_active = bool(active_matches)
        result["still_active"] = still_active
        if still_active:
            result["status_note"] = (
                "The API was restarted while this job was running. "
                "Phase shown below is from before the restart — live updates will resume shortly."
            )
            if not result.get("slurm_jobs"):
                try:
                    enriched = ssh_slurm.enrich_slurm_jobs_from_logs(
                        row["work_dir"], active_matches, conn,
                    )
                except Exception:
                    enriched = active_matches
                result["slurm_jobs"] = [
                    {
                        "job_id": m["job_id"],
                        "rule": m.get("rule"),
                        "status": m["state"],
                        "time": m.get("time", ""),
                        "genome": m.get("genome"),
                        "source": "fresh run",
                    }
                    for m in enriched
                ]
        elif _within_submit_grace(row):
            # No active SLURM jobs yet, but this row was touched moments ago:
            # it is still being submitted, not interrupted. Leave its status
            # alone -- marking it cancelled here is the very bug that made
            # not-yet-submitted jobs show up as cancelled.
            result["still_active"] = True
            result["status_note"] = (
                "This run is still being submitted to the cluster -- "
                "live updates will appear once its jobs are queued."
            )
        else:
            # Persisted status said non-terminal but there are no active jobs
            # and the row is old enough that submission would have surfaced
            # them by now. Normalize so this job no longer appears as running
            # forever.
            result["status"] = "cancelled"
            result["phase"] = "Interrupted (no active jobs)"
            try:
                job_history_client.record_job_updated(
                    conn,
                    main_db,
                    job_id,
                    status=result["status"],
                    phase=result["phase"],
                )
            except Exception as exc:
                LOGGER.warning("Could not persist stale-status correction for %s: %s", job_id, exc)

    if not result.get("logs") and row.get("work_dir"):
        try:
            result["logs"] = ssh_slurm.read_latest_snakemake_log(row["work_dir"], conn)
        except Exception:
            pass

    return result


def _reconcile_running(conn, main_db: str, rows: list[dict]) -> None:
    """Correct rows that claim to be running but are not.

    Workflow runs are detached (setsid + nohup) so they survive the GUI closing.
    The cost is that nothing writes a terminal status if the driver dies without
    reaching its exit sentinel -- a killed or wedged run then reads as "running"
    on the history page forever. One such row sat at "Submitting via SSH" after
    its driver was stopped.

    A run is considered alive if ANY of these hold:
      * its .rc sentinel is absent AND a driver process for it exists, or
      * it has SLURM jobs still queued/running.
    If the sentinel exists, the run finished and its exit code decides the
    status. If neither, it is interrupted.

    Deliberately read-mostly: only rows already marked running are touched, so a
    healthy run is never disturbed. Failures here are logged and ignored -- a
    history listing must not break because reconciliation could not run.
    """
    # EVERYTHING here is inside the try. This line was outside it, so an
    # unexpected row shape raised straight out of the function and 500'd
    # list_jobs -- which presented as an empty history page, i.e. reconciliation
    # destroyed the very listing it was meant to correct. A best-effort helper
    # must never be able to fail its caller.
    try:
        pending = [r for r in rows
                   if isinstance(r, dict) and (r.get("status") or "").lower() == "running"]
        if not pending:
            return
        ssh = conn.connect()
        for row in pending:
            jid = row.get("job_id") or row.get("id")
            if not jid:
                continue
            # Keep matching strict: substring matches produced false positives
            # and left finished jobs marked as running.
            escaped = re.escape(jid)
            probe = (
                f'rc=$(cat $HOME/.local/share/bsp/jobs/{jid}.rc 2>/dev/null); '
                f'drv=$(pgrep -u $USER -f "dane_wf.*{escaped}" 2>/dev/null | wc -l); '
                f'slurm=$(squeue -u $USER -h -o %j 2>/dev/null | grep -c "^{escaped}$"); '
                f'echo "${{rc:--}}|$drv|$slurm"'
            )
            _in, out, _err = ssh.exec_command(probe)
            reply = (out.read().decode() or "").strip().splitlines()
            if not reply:
                continue
            rc, drv, slurm = (reply[-1].split("|") + ["-", "0", "0"])[:3]
            try:
                drv_n, slurm_n = int(drv), int(slurm)
            except ValueError:
                continue

            if rc != "-":                      # finished: the sentinel decides
                ok = rc.strip() == "0"
                status, phase = ("completed", "Done") if ok else ("failed", f"Exited {rc.strip()}")
            elif drv_n > 0 or slurm_n > 0:     # genuinely still going
                continue
            elif _within_submit_grace(row):    # too young -- still submitting
                # A run mid-submission has no sentinel, no driver and no SLURM
                # jobs yet: identical to an interrupted run. Don't call a
                # freshly-touched row dead, or a job still on its way to the
                # cluster shows up as cancelled in history.
                continue
            else:                              # no sentinel, no driver, no jobs
                status, phase = "cancelled", "Interrupted (driver stopped)"

            LOGGER.info("reconciled job %s: running -> %s", jid, status)
            row["status"], row["phase"] = status, phase
            try:
                job_history_client.record_job_updated(
                    conn, main_db, jid, status=status, phase=phase)
            except Exception as exc:
                LOGGER.warning("could not persist reconciled status for %s: %s", jid, exc)
    except Exception as exc:
        LOGGER.warning("job reconciliation skipped: %s", exc)


@router.get("/jobs")
def list_jobs(
    workflow: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user),
):
    """List this user's persistent job history, optionally filtered to one
    workflow, paginated most-recent-first. A brand new user with no history
    at all gets an empty list, not an error."""
    empty_response = {"jobs": [], "page": page, "page_size": page_size, "total_jobs": 0, "total_pages": 1}

    conn = _build_connection(current_user)
    try:
        user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
    except Exception:
        return empty_response

    main_db = user_config.get('main_database')
    if not main_db:
        return empty_response

    offset = (page - 1) * page_size
    rows, total_jobs = job_history_client.list_jobs_and_count(
        conn,
        main_db,
        workflow=workflow,
        limit=page_size,
        offset=offset,
        owner_username=current_user["username"],
        owner_cluster_username=current_user["cluster_username"],
    )
    # Verify anything claiming to be running before reporting it as such.
    _reconcile_running(conn, main_db, rows)
    total_pages = max((total_jobs + page_size - 1) // page_size, 1)

    return {
        "jobs": [_job_from_history_row(row) for row in rows],
        "page": page,
        "page_size": page_size,
        "total_jobs": total_jobs,
        "total_pages": total_pages,
    }


@router.post("/cancel_job/{job_id}")
def cancel_job(job_id: str, current_user: dict = Depends(get_current_user)):
    """Emergency stop - cancel all SLURM jobs, kill remote process, and mark job as cancelled.

    Falls back to persistent history when the job isn't in the in-memory
    job_store -- e.g. after a dane-api restart -- the same fallback
    _resolve_job_work_dir already uses, so Emergency Stop still works for
    jobs the status page can still show via that fallback.
    """
    conn = _build_connection(current_user)
    job = job_store.get(job_id)
    main_db = None

    if job is not None:
        if job.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        slurm_ids = [sj["job_id"] for sj in job_store.get_slurm_jobs(job_id)]
    else:
        try:
            user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
        except Exception:
            raise HTTPException(status_code=404, detail="Job not found")

        main_db = user_config.get('main_database')
        row = (
            job_history_client.get_job(
                conn,
                main_db,
                job_id,
                owner_username=current_user["username"],
                owner_cluster_username=current_user["cluster_username"],
            )
            if main_db else None
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Job not found")
        slurm_ids = [sj["job_id"] for sj in (row.get("slurm_jobs") or [])]

    # The DRIVER goes first, and it is a SLURM job now rather than a login-node
    # process. Cancelling the children first would be pointless: Snakemake is
    # still alive at that moment and simply resubmits them. Killing the thing
    # that submits, then the things it submitted, is the only order that ends
    # the run.
    driver = None
    try:
        driver = ssh_slurm.driver_job_id(job_id, connection=conn)
    except Exception as exc:
        LOGGER.warning("Could not look up the driver job for %s: %s", job_id, exc)
    if driver:
        ssh_slurm.cancel_slurm_jobs([driver], connection=conn)
        LOGGER.info("Cancelled driver job %s for job %s", driver, job_id)

    # Cancel all SLURM subjobs
    if slurm_ids:
        ssh_slurm.cancel_slurm_jobs(slurm_ids, connection=conn)
        LOGGER.info("Cancelled %d SLURM jobs for job %s", len(slurm_ids), job_id)

    # Only for runs with no driver job -- ones started before the driver moved
    # into SLURM, which really do have a dane_wf on the login node. Skipped
    # otherwise because this is a pkill by name across the whole account: with
    # two runs in flight it would kill the other user-visible run too, and for
    # a driver on a compute node it cannot reach it anyway.
    if not driver:
        ssh_slurm.kill_remote_process("dane_wf", connection=conn)
        LOGGER.info("Killed remote dane_wf process for job %s", job_id)

    # Mark job as cancelled (this will also stop the status checker daemon, if any)
    if job is not None:
        job_store.cancel(job_id)
    else:
        job_history_client.record_job_updated(
            conn, main_db, job_id, status="cancelled", phase="Cancelled by user",
        )

    return {
        "success": True,
        "message": f"Cancelled job {job_id}",
        "slurm_jobs_cancelled": len(slurm_ids)
    }


def _main_db_for(current_user: dict, conn) -> tuple[str, dict]:
    """Fetch main_database from the user's config, raising HTTPException(400)
    if the config or that field is missing. Shared pre-flight for resume_job/
    restart_job (run_workflow does the same check inline as part of its
    richer missing_fields validation, which these two intentionally don't
    repeat in full -- they're relaunching an already-known-good prior
    submission, not validating a fresh one)."""
    try:
        user_config = ssh_sftp.read_remote_yaml(_config_path(current_user["home_dir"]), connection=conn)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Configuration file not found. Please create a configuration in your Profile settings first.",
        )
    main_db = _resolve_effective_main_db(current_user, conn, user_config, persist=True)
    return main_db, user_config


def _base_output_dir_from(path: str) -> str:
    """Strip a trailing /YYYY-MM-DD-HHMM timestamp segment off a previous
    job's work_dir/output_dir, recovering the base directory run_workflow
    originally built it from (see its own timestamp = ...strftime('%Y-%m-%d-%H%M'))."""
    return re.sub(r'/\d{4}-\d{2}-\d{2}-\d{4}$', '', path)


@router.post("/resume_job/{job_id}")
def resume_job(job_id: str, current_user: dict = Depends(get_current_user)):
    """Resume a failed (or stale-but-no-longer-active) job: copy its
    work_dir into a new timestamped folder, then relaunch the same
    genome_path/workflow/selected_tools there with margie_sb.resume: true
    so Snakemake's mtime-based rebuild skips whatever already completed."""
    conn = _build_connection(current_user)
    original = _load_job_for_action(job_id, current_user, conn)

    status = original.get("status")
    work_dir = original.get("work_dir")
    if status not in ("failed", "cancelled"):
        if status not in _POTENTIALLY_STALE_STATUSES:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot resume a job with status '{status}' -- only failed/cancelled jobs, "
                       "or non-terminal jobs no longer actually active on the cluster, can be resumed.",
            )
        # status looks non-terminal (pending/running/snakemake) -- this
        # could be a stale label left over from before a dane-api restart
        # (see get_job_status's still_active reconciliation), so actually
        # check the cluster before trusting it.
        still_active = True
        if work_dir:
            try:
                matches = ssh_slurm.find_active_jobs_in_workdir(
                    work_dir, current_user["cluster_username"], connection=conn,
                )
                still_active = any(m["state"] in ("RUNNING", "PENDING") for m in matches)
            except Exception as exc:
                LOGGER.warning("SLURM active-check failed while resuming job %s: %s", job_id, exc)
                still_active = True  # fail safe: don't resume something we couldn't confirm is dead
        if still_active:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot resume a job with status '{status}' while it appears to still be "
                       "active on the cluster.",
            )

    if not work_dir:
        raise HTTPException(status_code=400, detail="Original job has no recorded working directory to resume from")

    genome_path = original.get("genome_path")
    workflow = original.get("workflow")
    if not genome_path or not workflow:
        raise HTTPException(status_code=400, detail="Original job is missing genome_path/workflow, cannot resume")

    main_db, user_config = _main_db_for(current_user, conn)
    slurm_account = str(user_config.get('compute', {}).get('cluster_default', {}).get('account', '')).strip() or None
    slurm_partition = str(user_config.get('compute', {}).get('cluster_default', {}).get('partition', '')).strip() or None
    if not slurm_account:
        raise HTTPException(
            status_code=400,
            detail="compute.cluster_default.account is not configured. Please configure it in your Profile settings.",
        )
    _check_genome_path_exists(genome_path, workflow, conn)

    selected_tools_csv = original.get("selected_tools")
    selected_tools = selected_tools_csv.split(",") if selected_tools_csv else None

    return _launch_job(
        genome_path=genome_path, workflow=workflow,
        base_output_dir=_base_output_dir_from(work_dir), selected_tools=selected_tools,
        current_user=current_user, conn=conn, main_db=main_db,
        slurm_account=slurm_account, slurm_partition=slurm_partition,
        relaunched_from=job_id, copy_from_work_dir=work_dir,
    )


@router.post("/restart_job/{job_id}")
def restart_job(job_id: str, current_user: dict = Depends(get_current_user)):
    """Restart a job from scratch: same genome_path/workflow/selected_tools
    as a brand-new job/timestamp, with no copying -- functionally "Start
    New Analysis" auto-filled from a prior job. Works from any status."""
    conn = _build_connection(current_user)
    original = _load_job_for_action(job_id, current_user, conn)

    base_for_dir = original.get("work_dir") or original.get("output_dir")
    if not base_for_dir:
        raise HTTPException(status_code=400, detail="Original job has no recorded output directory to restart from")

    genome_path = original.get("genome_path")
    workflow = original.get("workflow")
    if not genome_path or not workflow:
        raise HTTPException(status_code=400, detail="Original job is missing genome_path/workflow, cannot restart")

    main_db, user_config = _main_db_for(current_user, conn)
    slurm_account = str(user_config.get('compute', {}).get('cluster_default', {}).get('account', '')).strip() or None
    slurm_partition = str(user_config.get('compute', {}).get('cluster_default', {}).get('partition', '')).strip() or None
    if not slurm_account:
        raise HTTPException(
            status_code=400,
            detail="compute.cluster_default.account is not configured. Please configure it in your Profile settings.",
        )
    _check_genome_path_exists(genome_path, workflow, conn)

    selected_tools_csv = original.get("selected_tools")
    selected_tools = selected_tools_csv.split(",") if selected_tools_csv else None

    return _launch_job(
        genome_path=genome_path, workflow=workflow,
        base_output_dir=_base_output_dir_from(base_for_dir), selected_tools=selected_tools,
        current_user=current_user, conn=conn, main_db=main_db,
        slurm_account=slurm_account, slurm_partition=slurm_partition,
        relaunched_from=job_id,
    )


@router.get("/job_files/{job_id}")
def get_job_files(
    job_id: str,
    subdir: str = "",
    current_user: dict = Depends(get_current_user),
):
    """List output files for a job via SFTP."""
    _validate_relative_path(subdir, label="subdirectory")

    conn = _build_connection(current_user)
    work_dir = _resolve_job_work_dir(job_id, current_user, conn)

    target_dir = f"{work_dir}/{subdir}".rstrip("/") if subdir else work_dir

    try:
        entries = ssh_sftp.list_remote_dir(target_dir, connection=conn)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Directory not found on remote")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list remote directory: {str(e)}")

    return {"work_dir": work_dir, "subdir": subdir, "entries": entries}


@router.get("/download_file/{job_id}")
def download_file(
    job_id: str,
    path: str,
    format: str = Query("raw", pattern="^(raw|excel)$"),
    current_user: dict = Depends(get_current_user),
):
    """Download a file from a job's working directory via SFTP.

    format=raw (default) streams the file unmodified. format=excel reads
    the whole delimited text file into memory and converts it to .xlsx
    before sending -- only sensible for the TSV/CSV outputs the viewer
    already supports, not arbitrary binary files.
    """
    _validate_relative_path(path)

    conn = _build_connection(current_user)
    work_dir = _resolve_job_work_dir(job_id, current_user, conn)

    remote_path = f"{work_dir}/{path}"
    filename = path.split("/")[-1]

    if format == "excel":
        xlsx_filename = re.sub(r"\.(tsv|csv)$", "", filename, flags=re.IGNORECASE) + ".xlsx"

        # Serve pre-generated .xlsx (from make-final-excel.py) when available.
        # Read eagerly so FileNotFoundError is caught here, not inside StreamingResponse.
        if re.search(r"\.(tsv|csv)$", path, re.IGNORECASE):
            xlsx_remote_path = re.sub(r"\.(tsv|csv)$", ".xlsx", remote_path, flags=re.IGNORECASE)
            try:
                xlsx_bytes = b"".join(ssh_sftp.stream_remote_file(xlsx_remote_path, connection=conn))
                return Response(
                    content=xlsx_bytes,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{xlsx_filename}"'},
                )
            except (FileNotFoundError, IOError):
                pass  # no pre-generated file — fall through to on-the-fly conversion
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")

        try:
            content = b"".join(ssh_sftp.stream_remote_file(remote_path, connection=conn))
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail="File not found on remote")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")

        text = content.decode("utf-8", errors="replace")
        delimiter = _detect_delimiter(path, text.splitlines()[0] if text else "")
        try:
            df = pd.read_csv(io.StringIO(text), sep=delimiter, dtype=str, keep_default_na=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse file as a table: {str(e)}")

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            _apply_tier_row_colors(writer.sheets["Sheet1"], df)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{xlsx_filename}"'},
        )

    try:
        return StreamingResponse(
            ssh_sftp.stream_remote_file(remote_path, connection=conn),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found on remote")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")


@router.get("/view_file/{job_id}")
def view_file(
    job_id: str,
    path: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    current_user: dict = Depends(get_current_user),
):
    """Read a paginated slice of a delimited text file from a job's
    working directory, for in-browser viewing without downloading the
    whole file. Cost scales with page position, not file size -- see
    ssh_sftp.read_remote_file_page."""
    _validate_relative_path(path)

    conn = _build_connection(current_user)
    work_dir = _resolve_job_work_dir(job_id, current_user, conn)

    remote_path = f"{work_dir}/{path}"

    cache_key = (job_id, path)
    known_total_lines = None
    try:
        mtime, size = ssh_sftp.stat_remote_file(remote_path, connection=conn)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found on remote")

    cached = _line_count_cache.get(cache_key)
    if cached and cached[0] == mtime and cached[1] == size:
        known_total_lines = cached[2]

    start_row = 2 + (page - 1) * page_size  # row 1 is always the header
    end_row = start_row + page_size - 1

    try:
        result = ssh_sftp.read_remote_file_page(
            remote_path, start_row, end_row, connection=conn,
            known_total_lines=known_total_lines,
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found on remote")
    except IsADirectoryError:
        raise HTTPException(status_code=400, detail="Path is a directory, not a file")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")

    total_lines = result["total_lines"]
    if known_total_lines is None:
        _line_count_cache[cache_key] = (mtime, size, total_lines)
    total_rows = max(total_lines - 1, 0)
    total_pages = max((total_rows + page_size - 1) // page_size, 1)

    delimiter = _detect_delimiter(path, result["header"])
    columns = result["header"].split(delimiter) if result["header"] else []
    rows = [line.split(delimiter) for line in result["lines"]]

    return {
        "columns": columns,
        "rows": rows,
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "total_pages": total_pages,
    }


@router.get("/job_status/{job_id}/stream")
def stream_job_status(job_id: str, current_user: dict = Depends(get_current_user)):
    """SSE endpoint that streams real-time job status updates."""
    job = job_store.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("user_id") != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    return StreamingResponse(
        job_runner.job_status_generator(job_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


@router.get("/all_genomes")
def all_genomes(path: str, current_user: dict = Depends(get_current_user)):
    """List genome files at a remote path on the user's cluster."""
    conn = _build_connection(current_user)
    genomes = ssh_slurm.get_genomes(path, connection=conn)
    return {"success": True, "Genomes": genomes}


def _resolve_browse_path(path: str, current_user: dict) -> str:
    """Expand a leading ~ to the user's home and normalize the path.

    Shared by /browse and /browse_view. There is deliberately no allowlist
    or traversal jail here: both endpoints act over the user's OWN SSH
    credentials, so the cluster's filesystem permissions are the access
    boundary -- a user who isn't in a depot's Unix group simply gets a
    permission error from the cluster (surfaced as 403 below), exactly as
    they would from a shell on the login node.
    """
    import posixpath

    if path.startswith("~"):
        path = path.replace("~", current_user["home_dir"], 1)
    return posixpath.normpath(path) if path else "/"


def _permission_status(exc: Exception) -> int:
    """Map a cluster filesystem error to 403 when it's a permission denial
    (paramiko surfaces these as OSError/IOError with errno EACCES, not
    always the PermissionError subclass), else 500."""
    import errno as _errno
    if isinstance(exc, PermissionError):
        return 403
    if isinstance(exc, OSError) and exc.errno == _errno.EACCES:
        return 403
    if "permission denied" in str(exc).lower():
        return 403
    return 500


@router.get("/browse")
def browse(path: str, current_user: dict = Depends(get_current_user)):
    """List a remote directory on the user's cluster for file-explorer
    navigation. Returns typed entries (directories first, then files, each
    alphabetical) plus the parent path so the front-end can render
    breadcrumbs and an "up" control.

    Listing runs over the user's own SSH credentials, so visibility is
    exactly what their cluster account can already see -- no extra
    traversal guard is needed here (unlike the job-relative endpoints,
    which restrict to a single job's work_dir).
    """
    import posixpath

    conn = _build_connection(current_user)
    path = _resolve_browse_path(path, current_user)

    try:
        kind = ssh_sftp.check_remote_path_kind(path, conn)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Path not found on cluster: '{path}'")
    except Exception as exc:
        raise HTTPException(status_code=_permission_status(exc),
                            detail=f"Could not access path: {exc}")

    if kind != "directory":
        raise HTTPException(status_code=400, detail=f"Not a directory: '{path}'")

    try:
        entries = ssh_sftp.list_remote_dir(path, connection=conn)
    except Exception as exc:
        status = _permission_status(exc)
        detail = f"Permission denied: '{path}'" if status == 403 else f"Failed to list directory: {exc}"
        raise HTTPException(status_code=status, detail=detail)

    entries.sort(key=lambda e: (e["type"] != "directory", e["name"].lower()))
    parent = posixpath.dirname(path.rstrip("/")) or "/"

    return {
        "success": True,
        "path": path,
        "parent": parent,
        "entries": entries,
    }


# Cap for the in-browser file viewer: read at most this many bytes so a huge
# output file can't blow up memory or the response. The UI flags truncation.
_VIEW_MAX_BYTES = 1_000_000


@router.get("/browse_view")
def browse_view(path: str, current_user: dict = Depends(get_current_user)):
    """Return the head of an arbitrary text file from the user's cluster for
    the file explorer's in-browser View button. Reads at most
    _VIEW_MAX_BYTES; binary files (detected via a NUL byte) are refused
    rather than dumped as mojibake. Same credential/permission model as
    /browse.
    """
    conn = _build_connection(current_user)
    path = _resolve_browse_path(path, current_user)

    try:
        kind = ssh_sftp.check_remote_path_kind(path, conn)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found on cluster: '{path}'")
    except Exception as exc:
        raise HTTPException(status_code=_permission_status(exc),
                            detail=f"Could not access path: {exc}")

    if kind == "directory":
        raise HTTPException(status_code=400, detail=f"Path is a directory, not a file: '{path}'")

    try:
        chunks, total = [], 0
        for chunk in ssh_sftp.stream_remote_file(path, connection=conn):
            chunks.append(chunk)
            total += len(chunk)
            if total >= _VIEW_MAX_BYTES:
                break
        raw = b"".join(chunks)[:_VIEW_MAX_BYTES]
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found on remote")
    except Exception as exc:
        status = _permission_status(exc)
        detail = f"Permission denied: '{path}'" if status == 403 else f"Failed to read file: {exc}"
        raise HTTPException(status_code=status, detail=detail)

    if b"\x00" in raw:
        return {
            "path": path,
            "binary": True,
            "truncated": False,
            "content": "",
        }

    return {
        "path": path,
        "binary": False,
        "truncated": total >= _VIEW_MAX_BYTES,
        "content": raw.decode("utf-8", errors="replace"),
    }


@router.post("/browse_save")
def browse_save(payload: dict, current_user: dict = Depends(get_current_user)):
    """Write edited text back to a file on the user's cluster from the file
    explorer's in-browser editor. Refuses to save over a directory, and
    refuses truncated content (the viewer only loaded part of a large file,
    so saving it would silently discard the rest). Same credential/permission
    model as /browse and /browse_view.
    """
    path = payload.get("path", "").strip()
    content = payload.get("content", "")
    if not path:
        raise HTTPException(status_code=400, detail="Path is required")
    if payload.get("truncated"):
        raise HTTPException(status_code=400, detail="File is too large to edit in-browser")

    conn = _build_connection(current_user)
    resolved_path = _resolve_browse_path(path, current_user)

    try:
        kind = ssh_sftp.check_remote_path_kind(resolved_path, conn)
        if kind == "directory":
            raise HTTPException(status_code=400, detail=f"Path is a directory, not a file: '{resolved_path}'")
    except FileNotFoundError:
        pass  # new file — fine to create on save

    try:
        ssh_sftp.write_remote_text_file(resolved_path, content, connection=conn)
    except Exception as exc:
        status = _permission_status(exc)
        detail = f"Permission denied: '{resolved_path}'" if status == 403 else f"Failed to save file: {exc}"
        raise HTTPException(status_code=status, detail=detail)

    return {"success": True, "path": resolved_path}
